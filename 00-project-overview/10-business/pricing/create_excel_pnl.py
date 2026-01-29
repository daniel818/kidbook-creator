#!/usr/bin/env python3
"""
KidBook Creator - P&L Forecasting Excel Generator with Formatting
Creates a professional Excel workbook with multiple sheets and formatting
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_pnl_excel():
    """Create a comprehensive P&L Excel workbook"""
    
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    currency_format = '$#,##0.00'
    percent_format = '0.0%'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============================================
    # SHEET 1: ASSUMPTIONS
    # ============================================
    ws1 = wb.active
    ws1.title = "Assumptions"
    
    ws1['A1'] = "KIDBOOK CREATOR - P&L FORECASTING MODEL"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = f"Version 2.0 | Last Updated: {datetime.now().strftime('%Y-%m-%d')}"
    
    row = 4
    ws1[f'A{row}'] = "FIXED MONTHLY COSTS"
    ws1[f'A{row}'].font = subheader_font
    ws1[f'A{row}'].fill = subheader_fill
    row += 1
    
    fixed_costs = [
        ("Supabase Pro", 25),
        ("Netlify", 9),
        ("Buffer for Additional Services", 100),
    ]
    
    for item, cost in fixed_costs:
        ws1[f'A{row}'] = item
        ws1[f'B{row}'] = cost
        ws1[f'B{row}'].number_format = currency_format
        row += 1
    
    ws1[f'A{row}'] = "TOTAL FIXED MONTHLY"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'] = 134
    ws1[f'B{row}'].number_format = currency_format
    ws1[f'B{row}'].font = Font(bold=True)
    
    row += 2
    ws1[f'A{row}'] = "ONE-TIME SETUP COSTS"
    ws1[f'A{row}'].font = subheader_font
    ws1[f'A{row}'].fill = subheader_fill
    row += 1
    
    setup_costs = [
        ("Gemini API Credits", 150),
        ("Domain Registration", 100),
        ("Example Books Creation", 75),
        ("Friends & Family Testing", 350),
    ]
    
    for item, cost in setup_costs:
        ws1[f'A{row}'] = item
        ws1[f'B{row}'] = cost
        ws1[f'B{row}'].number_format = currency_format
        row += 1
    
    ws1[f'A{row}'] = "TOTAL SETUP COSTS"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'] = 675
    ws1[f'B{row}'].number_format = currency_format
    ws1[f'B{row}'].font = Font(bold=True)
    
    row += 2
    ws1[f'A{row}'] = "PRICING"
    ws1[f'A{row}'].font = subheader_font
    ws1[f'A{row}'].fill = subheader_fill
    row += 1
    ws1[f'A{row}'] = "Digital Book (PDF)"
    ws1[f'B{row}'] = 15
    ws1[f'B{row}'].number_format = currency_format
    row += 1
    ws1[f'A{row}'] = "Printed Hardcover Book"
    ws1[f'B{row}'] = 45
    ws1[f'B{row}'].number_format = currency_format
    
    row += 2
    ws1[f'A{row}'] = "CONVERSION FUNNEL (EDITABLE)"
    ws1[f'A{row}'].font = subheader_font
    ws1[f'A{row}'].fill = subheader_fill
    row += 1
    digital_conv_row = row
    ws1[f'A{row}'] = "Preview → Digital Conversion"
    ws1[f'B{row}'] = 0.02
    ws1[f'B{row}'].number_format = percent_format
    ws1[f'C{row}'] = "← Edit this"
    row += 1
    printed_conv_row = row
    ws1[f'A{row}'] = "Preview → Printed Conversion"
    ws1[f'B{row}'] = 0.02
    ws1[f'B{row}'].number_format = percent_format
    ws1[f'C{row}'] = "← Edit this"
    
    row += 2
    ws1[f'A{row}'] = "VARIABLE COSTS - PREVIEW"
    ws1[f'A{row}'].font = subheader_font
    ws1[f'A{row}'].fill = subheader_fill
    row += 1
    ws1[f'A{row}'] = "AI Generation (1 image + text)"
    ws1[f'B{row}'] = 0.09
    ws1[f'B{row}'].number_format = currency_format
    
    row += 2
    ws1[f'A{row}'] = "VARIABLE COSTS - DIGITAL BOOK"
    ws1[f'A{row}'].font = subheader_font
    ws1[f'A{row}'].fill = subheader_fill
    row += 1
    digital_costs = [
        ("AI Generation", 0.50),
        ("Storage (permanent)", 0.02),
        ("Payment Processing (2.9% + $0.30)", 0.74),
    ]
    for item, cost in digital_costs:
        ws1[f'A{row}'] = item
        ws1[f'B{row}'] = cost
        ws1[f'B{row}'].number_format = currency_format
        row += 1
    ws1[f'A{row}'] = "TOTAL COST PER DIGITAL"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'] = 1.26
    ws1[f'B{row}'].number_format = currency_format
    ws1[f'B{row}'].font = Font(bold=True)
    
    row += 2
    ws1[f'A{row}'] = "VARIABLE COSTS - PRINTED HARDCOVER"
    ws1[f'A{row}'].font = subheader_font
    ws1[f'A{row}'].fill = subheader_fill
    row += 1
    printed_costs = [
        ("AI Generation", 0.50),
        ("Storage (permanent)", 0.02),
        ("Printing (Lulu - 24 pages)", 14.50),
        ("Shipping (EU average)", 12.00),
        ("Payment Processing (2.9% + $0.30)", 1.61),
    ]
    for item, cost in printed_costs:
        ws1[f'A{row}'] = item
        ws1[f'B{row}'] = cost
        ws1[f'B{row}'].number_format = currency_format
        row += 1
    ws1[f'A{row}'] = "TOTAL COST PER PRINTED"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'] = 28.63
    ws1[f'B{row}'].number_format = currency_format
    ws1[f'B{row}'].font = Font(bold=True)
    
    # Set column widths
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 20
    
    # ============================================
    # SHEET 2: MONTHLY P&L
    # ============================================
    ws2 = wb.create_sheet("Monthly P&L")
    
    ws2['A1'] = "KIDBOOK CREATOR - 12 MONTH P&L PROJECTION"
    ws2['A1'].font = Font(bold=True, size=14)
    
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    preview_volumes = [50, 75, 100, 150, 200, 250, 300, 350, 400, 450, 500, 550]
    
    # Headers
    row = 3
    ws2[f'A{row}'] = "MONTH"
    ws2[f'A{row}'].font = header_font
    ws2[f'A{row}'].fill = header_fill
    for i, month in enumerate(months):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = month
        ws2[f'{col}{row}'].font = header_font
        ws2[f'{col}{row}'].fill = header_fill
        ws2[f'{col}{row}'].alignment = Alignment(horizontal='center')
    ws2[f'N{row}'] = "TOTAL"
    ws2[f'N{row}'].font = header_font
    ws2[f'N{row}'].fill = header_fill
    
    row += 2
    ws2[f'A{row}'] = "VOLUME METRICS (EDITABLE)"
    ws2[f'A{row}'].font = subheader_font
    ws2[f'A{row}'].fill = subheader_fill
    
    row += 1
    ws2[f'A{row}'] = "Preview Generations"
    for i, vol in enumerate(preview_volumes):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = vol
        ws2[f'{col}{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].font = Font(bold=True)
    
    # Conversions
    row += 1
    digital_conv_calc_row = row
    ws2[f'A{row}'] = "Digital Conversions (2%)"
    for i in range(12):
        col = get_column_letter(i + 2)
        prev_col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={prev_col}{row-1}*Assumptions!$B${digital_conv_row}"
        ws2[f'{col}{row}'].number_format = '0.0'
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    
    row += 1
    printed_conv_calc_row = row
    ws2[f'A{row}'] = "Printed Conversions (2%)"
    for i in range(12):
        col = get_column_letter(i + 2)
        prev_col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={prev_col}{row-2}*Assumptions!$B${printed_conv_row}"
        ws2[f'{col}{row}'].number_format = '0.0'
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    
    # Revenue
    row += 2
    ws2[f'A{row}'] = "REVENUE"
    ws2[f'A{row}'].font = subheader_font
    ws2[f'A{row}'].fill = subheader_fill
    
    row += 1
    digital_rev_row = row
    ws2[f'A{row}'] = "Digital Books ($15 each)"
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={col}{digital_conv_calc_row}*Assumptions!$B$16"
        ws2[f'{col}{row}'].number_format = currency_format
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    
    row += 1
    printed_rev_row = row
    ws2[f'A{row}'] = "Printed Books ($45 each)"
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={col}{printed_conv_calc_row}*Assumptions!$B$17"
        ws2[f'{col}{row}'].number_format = currency_format
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    
    row += 1
    total_rev_row = row
    ws2[f'A{row}'] = "TOTAL REVENUE"
    ws2[f'A{row}'].font = Font(bold=True)
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={col}{digital_rev_row}+{col}{printed_rev_row}"
        ws2[f'{col}{row}'].number_format = currency_format
        ws2[f'{col}{row}'].font = Font(bold=True)
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    ws2[f'N{row}'].font = Font(bold=True)
    
    # Variable Costs
    row += 2
    ws2[f'A{row}'] = "VARIABLE COSTS"
    ws2[f'A{row}'].font = subheader_font
    ws2[f'A{row}'].fill = subheader_fill
    
    row += 1
    preview_cost_row = row
    ws2[f'A{row}'] = "Preview Generation Costs"
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={col}6*Assumptions!$B$24"
        ws2[f'{col}{row}'].number_format = currency_format
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    
    row += 1
    digital_cost_row = row
    ws2[f'A{row}'] = "Digital Book Costs"
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={col}{digital_conv_calc_row}*Assumptions!$B$32"
        ws2[f'{col}{row}'].number_format = currency_format
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    
    row += 1
    printed_cost_row = row
    ws2[f'A{row}'] = "Printed Book Costs"
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={col}{printed_conv_calc_row}*Assumptions!$B$42"
        ws2[f'{col}{row}'].number_format = currency_format
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    
    row += 1
    refund_row = row
    ws2[f'A{row}'] = "Refunds (3% of revenue)"
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={col}{total_rev_row}*0.03"
        ws2[f'{col}{row}'].number_format = currency_format
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    
    row += 1
    failed_row = row
    ws2[f'A{row}'] = "Failed Prints (1.5%)"
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={col}{printed_conv_calc_row}*Assumptions!$B$42*0.015"
        ws2[f'{col}{row}'].number_format = currency_format
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    
    row += 1
    total_var_row = row
    ws2[f'A{row}'] = "TOTAL VARIABLE COSTS"
    ws2[f'A{row}'].font = Font(bold=True)
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"=SUM({col}{preview_cost_row}:{col}{failed_row})"
        ws2[f'{col}{row}'].number_format = currency_format
        ws2[f'{col}{row}'].font = Font(bold=True)
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    ws2[f'N{row}'].font = Font(bold=True)
    
    # Gross Profit
    row += 2
    gross_profit_row = row
    ws2[f'A{row}'] = "GROSS PROFIT"
    ws2[f'A{row}'].font = Font(bold=True)
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={col}{total_rev_row}-{col}{total_var_row}"
        ws2[f'{col}{row}'].number_format = currency_format
        ws2[f'{col}{row}'].font = Font(bold=True)
        ws2[f'{col}{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    ws2[f'N{row}'].font = Font(bold=True)
    ws2[f'N{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    row += 1
    ws2[f'A{row}'] = "Gross Margin %"
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"=IF({col}{total_rev_row}>0,{col}{gross_profit_row}/{col}{total_rev_row},0)"
        ws2[f'{col}{row}'].number_format = percent_format
    ws2[f'N{row}'] = f"=IF(N{total_rev_row}>0,N{gross_profit_row}/N{total_rev_row},0)"
    ws2[f'N{row}'].number_format = percent_format
    
    # Fixed Costs
    row += 2
    ws2[f'A{row}'] = "FIXED COSTS"
    ws2[f'A{row}'].font = subheader_font
    ws2[f'A{row}'].fill = subheader_fill
    
    row += 1
    fixed_row = row
    ws2[f'A{row}'] = "Monthly Fixed Costs"
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = "=Assumptions!$B$9"
        ws2[f'{col}{row}'].number_format = currency_format
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    
    # Net Profit
    row += 2
    net_profit_row = row
    ws2[f'A{row}'] = "NET PROFIT"
    ws2[f'A{row}'].font = Font(bold=True, size=12)
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"={col}{gross_profit_row}-{col}{fixed_row}"
        ws2[f'{col}{row}'].number_format = currency_format
        ws2[f'{col}{row}'].font = Font(bold=True)
        ws2[f'{col}{row}'].fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    ws2[f'N{row}'] = f"=SUM(B{row}:M{row})"
    ws2[f'N{row}'].number_format = currency_format
    ws2[f'N{row}'].font = Font(bold=True)
    ws2[f'N{row}'].fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    
    row += 1
    ws2[f'A{row}'] = "Net Margin %"
    for i in range(12):
        col = get_column_letter(i + 2)
        ws2[f'{col}{row}'] = f"=IF({col}{total_rev_row}>0,{col}{net_profit_row}/{col}{total_rev_row},0)"
        ws2[f'{col}{row}'].number_format = percent_format
    ws2[f'N{row}'] = f"=IF(N{total_rev_row}>0,N{net_profit_row}/N{total_rev_row},0)"
    ws2[f'N{row}'].number_format = percent_format
    
    # Set column widths
    ws2.column_dimensions['A'].width = 30
    for i in range(2, 15):
        ws2.column_dimensions[get_column_letter(i)].width = 12
    
    # ============================================
    # SHEET 3: SENSITIVITY ANALYSIS
    # ============================================
    ws3 = wb.create_sheet("Sensitivity Analysis")
    
    ws3['A1'] = "SENSITIVITY ANALYSIS"
    ws3['A1'].font = Font(bold=True, size=14)
    
    row = 3
    ws3[f'A{row}'] = "CONVERSION RATE IMPACT"
    ws3[f'A{row}'].font = subheader_font
    ws3[f'A{row}'].fill = subheader_fill
    ws3[f'A{row+1}'] = "(Monthly: 300 Previews)"
    
    row += 3
    ws3[f'A{row}'] = "Conversion Rate"
    rates = [0.02, 0.03, 0.04, 0.05, 0.06]
    for i, rate in enumerate(rates):
        col = get_column_letter(i + 2)
        ws3[f'{col}{row}'] = rate
        ws3[f'{col}{row}'].number_format = percent_format
        ws3[f'{col}{row}'].font = Font(bold=True)
        ws3[f'{col}{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row += 1
    ws3[f'A{row}'] = "Total Conversions"
    for i, rate in enumerate(rates):
        col = get_column_letter(i + 2)
        ws3[f'{col}{row}'] = 300 * rate
        ws3[f'{col}{row}'].number_format = '0.0'
    
    row += 1
    ws3[f'A{row}'] = "Revenue"
    for i, rate in enumerate(rates):
        col = get_column_letter(i + 2)
        conversions = 300 * rate
        digital = conversions * 0.25 * 15
        printed = conversions * 0.75 * 45
        ws3[f'{col}{row}'] = digital + printed
        ws3[f'{col}{row}'].number_format = currency_format
    
    row += 1
    ws3[f'A{row}'] = "Variable Costs"
    for i, rate in enumerate(rates):
        col = get_column_letter(i + 2)
        conversions = 300 * rate
        preview_cost = 300 * 0.09
        digital_cost = conversions * 0.25 * 1.26
        printed_cost = conversions * 0.75 * 28.63
        ws3[f'{col}{row}'] = preview_cost + digital_cost + printed_cost
        ws3[f'{col}{row}'].number_format = currency_format
    
    row += 1
    ws3[f'A{row}'] = "Net Profit"
    ws3[f'A{row}'].font = Font(bold=True)
    for i in range(5):
        col = get_column_letter(i + 2)
        ws3[f'{col}{row}'] = f"={col}{row-2}-{col}{row-1}-134"
        ws3[f'{col}{row}'].number_format = currency_format
        ws3[f'{col}{row}'].font = Font(bold=True)
    
    row += 3
    ws3[f'A{row}'] = "PRICING IMPACT"
    ws3[f'A{row}'].font = subheader_font
    ws3[f'A{row}'].fill = subheader_fill
    ws3[f'A{row+1}'] = "(Monthly: 300 Previews, 4% Conversion)"
    
    row += 3
    ws3[f'A{row}'] = "Printed Book Price"
    prices = [40, 42, 45, 48, 50]
    for i, price in enumerate(prices):
        col = get_column_letter(i + 2)
        ws3[f'{col}{row}'] = price
        ws3[f'{col}{row}'].number_format = currency_format
        ws3[f'{col}{row}'].font = Font(bold=True)
        ws3[f'{col}{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row += 1
    ws3[f'A{row}'] = "Total Revenue"
    for i, price in enumerate(prices):
        col = get_column_letter(i + 2)
        digital_rev = 300 * 0.04 * 0.25 * 15
        printed_rev = 300 * 0.04 * 0.75 * price
        ws3[f'{col}{row}'] = digital_rev + printed_rev
        ws3[f'{col}{row}'].number_format = currency_format
    
    row += 1
    ws3[f'A{row}'] = "Net Profit"
    ws3[f'A{row}'].font = Font(bold=True)
    for i in range(5):
        col = get_column_letter(i + 2)
        # Fixed variable costs at 4% conversion
        var_costs = 300 * 0.09 + 300 * 0.04 * 0.25 * 1.26 + 300 * 0.04 * 0.75 * 28.63
        ws3[f'{col}{row}'] = f"={col}{row-1}-{var_costs}-134"
        ws3[f'{col}{row}'].number_format = currency_format
        ws3[f'{col}{row}'].font = Font(bold=True)
    
    # Set column widths
    ws3.column_dimensions['A'].width = 30
    for i in range(2, 8):
        ws3.column_dimensions[get_column_letter(i)].width = 15
    
    # ============================================
    # SHEET 4: SUMMARY DASHBOARD
    # ============================================
    ws4 = wb.create_sheet("Summary Dashboard")
    
    ws4['A1'] = "KIDBOOK CREATOR - EXECUTIVE SUMMARY"
    ws4['A1'].font = Font(bold=True, size=14)
    
    row = 3
    ws4[f'A{row}'] = "KEY METRICS (Year 1)"
    ws4[f'A{row}'].font = subheader_font
    ws4[f'A{row}'].fill = subheader_fill
    
    row += 2
    ws4[f'A{row}'] = "Total Preview Generations"
    ws4[f'B{row}'] = "='Monthly P&L'!N6"
    ws4[f'B{row}'].font = Font(bold=True, size=12)
    
    row += 1
    ws4[f'A{row}'] = "Total Conversions"
    ws4[f'B{row}'] = f"='Monthly P&L'!N{digital_conv_calc_row}+'Monthly P&L'!N{printed_conv_calc_row}"
    ws4[f'B{row}'].font = Font(bold=True, size=12)
    ws4[f'B{row}'].number_format = '0.0'
    
    row += 2
    ws4[f'A{row}'] = "FINANCIAL PERFORMANCE"
    ws4[f'A{row}'].font = subheader_font
    ws4[f'A{row}'].fill = subheader_fill
    
    row += 1
    ws4[f'A{row}'] = "Total Revenue"
    ws4[f'B{row}'] = f"='Monthly P&L'!N{total_rev_row}"
    ws4[f'B{row}'].number_format = currency_format
    ws4[f'B{row}'].font = Font(bold=True, size=12)
    
    row += 1
    ws4[f'A{row}'] = "Gross Profit"
    ws4[f'B{row}'] = f"='Monthly P&L'!N{gross_profit_row}"
    ws4[f'B{row}'].number_format = currency_format
    ws4[f'B{row}'].font = Font(bold=True, size=12)
    
    row += 1
    ws4[f'A{row}'] = "Net Profit"
    ws4[f'B{row}'] = f"='Monthly P&L'!N{net_profit_row}"
    ws4[f'B{row}'].number_format = currency_format
    ws4[f'B{row}'].font = Font(bold=True, size=12)
    ws4[f'B{row}'].fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    
    row += 2
    ws4[f'A{row}'] = "UNIT ECONOMICS"
    ws4[f'A{row}'].font = subheader_font
    ws4[f'A{row}'].fill = subheader_fill
    
    row += 2
    ws4[f'A{row}'] = "Per Digital Book:"
    ws4[f'A{row}'].font = Font(bold=True)
    row += 1
    ws4[f'A{row}'] = "  Revenue"
    ws4[f'B{row}'] = 15
    ws4[f'B{row}'].number_format = currency_format
    row += 1
    ws4[f'A{row}'] = "  Cost"
    ws4[f'B{row}'] = 1.26
    ws4[f'B{row}'].number_format = currency_format
    row += 1
    ws4[f'A{row}'] = "  Gross Profit"
    ws4[f'B{row}'] = 13.74
    ws4[f'B{row}'].number_format = currency_format
    ws4[f'B{row}'].font = Font(bold=True)
    ws4[f'C{row}'] = "91.6%"
    
    row += 2
    ws4[f'A{row}'] = "Per Printed Book:"
    ws4[f'A{row}'].font = Font(bold=True)
    row += 1
    ws4[f'A{row}'] = "  Revenue"
    ws4[f'B{row}'] = 45
    ws4[f'B{row}'].number_format = currency_format
    row += 1
    ws4[f'A{row}'] = "  Cost"
    ws4[f'B{row}'] = 28.63
    ws4[f'B{row}'].number_format = currency_format
    row += 1
    ws4[f'A{row}'] = "  Gross Profit"
    ws4[f'B{row}'] = 16.37
    ws4[f'B{row}'].number_format = currency_format
    ws4[f'B{row}'].font = Font(bold=True)
    ws4[f'C{row}'] = "36.4%"
    
    row += 2
    ws4[f'A{row}'] = "BREAK-EVEN ANALYSIS"
    ws4[f'A{row}'].font = subheader_font
    ws4[f'A{row}'].fill = subheader_fill
    
    row += 1
    ws4[f'A{row}'] = "Monthly Fixed Costs"
    ws4[f'B{row}'] = 134
    ws4[f'B{row}'].number_format = currency_format
    
    row += 1
    ws4[f'A{row}'] = "Previews needed per month"
    ws4[f'B{row}'] = 262
    ws4[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws4[f'A{row}'] = "Conversions needed per month"
    ws4[f'B{row}'] = 10.5
    ws4[f'B{row}'].font = Font(bold=True)
    ws4[f'B{row}'].number_format = '0.0'
    
    # Set column widths
    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 15
    
    # Save workbook
    filename = "KidBook_Creator_PNL_Forecast.xlsx"
    wb.save(filename)
    print(f"✓ Excel file created: {filename}")
    print(f"\nThe workbook contains 4 sheets:")
    print("  1. Assumptions - All cost and pricing parameters")
    print("  2. Monthly P&L - 12-month projection (edit preview volumes)")
    print("  3. Sensitivity Analysis - Conversion rate and pricing scenarios")
    print("  4. Summary Dashboard - Key metrics and unit economics")
    print(f"\nFile location: {filename}")

if __name__ == "__main__":
    create_pnl_excel()
